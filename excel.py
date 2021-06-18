from openpyxl import load_workbook
from datetime import timedelta,date
from re import search

wb = load_workbook('load/hts_rescreen_dates.xlsx')
ws = wb.active

data = [
    "2020-11-23",
    "2020-11-24",
    "2020-11-25",
    "2020-11-26",
    "2020-11-27",
    "2020-11-30",
    "2020-12-01",
    "2020-12-02",
    "2020-12-03",
    "2020-12-04",
    "2020-12-07",
    "2020-12-08",
    "2020-12-09",
    "2020-12-10",
    "2020-12-11",
    "2020-12-14",
    "2020-12-15",
    "2020-12-16",
    "2020-12-17",
    "2020-12-18",
    "2020-12-21",
    "2020-12-22",
    "2020-12-23",
    "2020-12-28",
    "2020-12-29",
    "2020-12-30",
    "2021-01-04",
    "2021-01-05",
]


headings = ['Dates','Rescreening 1','Rescreening 1 With Days','Rescreening 2','Rescreening 2 With Days']
for c in range(1,len(headings)+1):
    ws.cell(row=1,column=c,value= headings[c-1])
for row in data:
    i = data.index(row)
    day11 = (date.fromisoformat(row) + timedelta(days=95))
    day12 = date.ctime(day11)
    day21 = (date.fromisoformat(row) + timedelta(days=98))
    day22 = date.ctime(day21)

    day31 = (date.fromisoformat(row) + timedelta(days=93*2))
    day32 = date.ctime(day31)

    day41 = (date.fromisoformat(row) + timedelta(days=(93*2)+3))
    day42 = date.ctime(day41)
    ws.cell(row=(i+2),column=1,value=row)
    ws.cell(row=(i+2),column=2,value=day11)
    ws.cell(row=(i+2),column=3,value=day12)
    ws.cell(row=(i+2),column=4,value=day31)
    ws.cell(row=(i+2),column=5,value=day32)
    # remove weekends
    if search('Sat',day12) or search('Sun',day12):
        ws.cell(row=(i+2),column=2,value=day21)
        ws.cell(row=(i+2),column=3,value=day22)
    if search('Sat',day32) or search('Sun',day32):
        ws.cell(row=(i+2),column=4,value=day41)
        ws.cell(row=(i+2),column=5,value=day42)
    


wb.save('load/hts_rescreen_dates.xlsx')
